import keyboard
import pyperclip
import pandas as pd
import pyautogui
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

# Define a dictionary to store key combinations
keybindings = {
    # Resulting comments scripts
    'Ctrl+shift+1': r'C:\Users\Kliea\Documents\Dev\Proj\Python\AutoKB\Scripts\DNAI Cultural Buisiness.txt',
    'Ctrl+shift+2': r'C:\Users\Kliea\Documents\Dev\Proj\Python\AutoKB\Scripts\DNAI Sorry Buisiness.txt',
    'Ctrl+shift+3': r'C:\Users\Kliea\Documents\Dev\Proj\Python\AutoKB\Scripts\DNAV community Unrest.txt',
    'Ctrl+shift+4': r'C:\Users\Kliea\Documents\Dev\Proj\Python\AutoKB\Scripts\DNAV Medical.txt',
    'Ctrl+shift+5': r'C:\Users\Kliea\Documents\Dev\Proj\Python\AutoKB\Scripts\DNAV Non-compelable.txt',
    'Ctrl+shift+6': r'C:\Users\Kliea\Documents\Dev\Proj\Python\AutoKB\Scripts\DNAV Work related.txt',
    'Ctrl+shift+7': r'C:\Users\Kliea\Documents\Dev\Proj\Python\AutoKB\Scripts\DNAV VOLUNTARY.txt',
    'Ctrl+shift+8': r'C:\Users\Kliea\Documents\Dev\Proj\Python\AutoKB\Scripts\DNAV NO STAFF IN COMMUNITY.txt',
    'Ctrl+shift+9': r'C:\Users\Kliea\Documents\Dev\Proj\Python\AutoKB\Scripts\DNAV FLOODING.txt',
    'Ctrl+shift+0': r'C:\Users\Kliea\Documents\Dev\Proj\Python\AutoKB\Scripts\DNAI CONTACTABLE.txt',
    'Ctrl+shift+-': r'C:\Users\Kliea\Documents\Dev\Proj\Python\AutoKB\Scripts\DNAI noncontactable.txt',
    'Ctrl+shift+=': r'C:\Users\Kliea\Documents\Dev\Proj\Python\AutoKB\Scripts\DNAI INCOME PHONE.txt',
    'Ctrl+shift+`': r'C:\Users\Kliea\Documents\Dev\Proj\Python\AutoKB\Scripts\DNAI INCOME.txt',
    #comment scripts keybinds
    'ctrl+1': r'C:\Users\Kliea\Documents\Dev\Proj\Python\AutoKB\Scripts\DNAI Cultural Buisiness.txt',
    'ctrl+2': r'C:\Users\Kliea\Documents\Dev\Proj\Python\AutoKB\Scripts\DNAI Sorry Buisiness.txt',
    'ctrl+3': r'C:\Users\Kliea\Documents\Dev\Proj\Python\AutoKB\Scripts\DNAV community Unrest.txt',
    'ctrl+4': r'C:\Users\Kliea\Documents\Dev\Proj\Python\AutoKB\Scripts\DNAV Medical.txt',
    'ctrl+5': r'C:\Users\Kliea\Documents\Dev\Proj\Python\AutoKB\Scripts\DNAV Non-compelable.txt',
    'ctrl+6': r'C:\Users\Kliea\Documents\Dev\Proj\Python\AutoKB\Scripts\DNAV Work related.txt',
    'ctrl+7': r'C:\Users\Kliea\Documents\Dev\Proj\Python\AutoKB\Scripts\DNAV VOLUNTARY.txt',
    'ctrl+8': r'C:\Users\Kliea\Documents\Dev\Proj\Python\AutoKB\Scripts\DNAV NO STAFF IN COMMUNITY.txt',
    'ctrl+9': r'C:\Users\Kliea\Documents\Dev\Proj\Python\AutoKB\Scripts\DNAV FLOODING.txt',
    'ctrl+0': r'C:\Users\Kliea\Documents\Dev\Proj\Python\AutoKB\Scripts\DNAI CONTACTABLE.txt',
    'ctrl+-': r'C:\Users\Kliea\Documents\Dev\Proj\Python\AutoKB\Scripts\DNAI noncontactable.txt',
    'ctrl+=': r'C:\Users\Kliea\Documents\Dev\Proj\Python\AutoKB\Scripts\DNAI INCOME PHONE.txt',
    'ctrl+shift+1': r'C:\Users\Kliea\Documents\Dev\Proj\Python\AutoKB\Scripts\DNAI INCOME.txt',
    # Add other keybindings
    'ctrl+shift+alt+0': r'C:\Users\Kliea\Documents\Dev\Proj\Python\AutoKB\Reports\216-CDPJobPlanMonitoring-20240428.csv',
    # Appt Comment Script
    'ctrl+ALT+SHIFT+1': r'C:\Users\Kliea\Documents\Dev\Proj\Python\AutoKB\Scripts\appt script re-engagements.txt'
}

def paste_text(filepath):
    # Open the text document and copy the content to the clipboard
    with open(filepath, 'r') as file:
        content = file.read().strip()  # Strip trailing newline characters
    pyperclip.copy(content)  # Copy content to clipboard
    pyautogui.hotkey('ctrl', 'v')  # Paste content from clipboard

def process_csv_and_save_as_xlsx(file_path, columns_to_delete, output_file):
    df = pd.read_csv(file_path)

    # Delete specified columns
    df.drop(columns=df.columns[columns_to_delete], inplace=True)

    # Create a workbook
    wb = Workbook()
    ws = wb.active

    # Convert DataFrame to Excel
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Apply thin borders to all cells
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    # Apply thick outside border to the first row
    thick_border = Border(top=Side(style='thick'), 
                          left=Side(style='thick'), 
                          right=Side(style='thick'), 
                          bottom=Side(style='thick'))
    for cell in ws[1]:
        cell.border = thick_border

    # Apply blue background and bold text to the first row
    for cell in ws[1]:
        cell.fill = PatternFill(start_color='3366FF', end_color='3366FF', fill_type='solid')
        cell.font = Font(bold=True)

    # Save the workbook
    wb.save(output_file)

def on_hotkey(hotkey):
    # Get the file path associated with the hotkey
    filepath = keybindings.get(hotkey)
    if filepath:
        if filepath.endswith('.txt'):
            # Paste content from the txt doc into the active window
            paste_text(filepath)
        elif filepath.endswith('.csv'):
            # Process the CSV file and save as XLSX
            columns_to_delete = [0, 1, 2, 3, 4, 5, 6, 8, 12, 15, 16, 17, 18, 81, 19,
                                20, 25, 27, 29, 30, 33, 34, 42, 44, 45, 47, 48, 50,
                                51, 52, 53, 54, 55, 56, 57, 58, 59, 60, 62, 63,
                                64, 65, 66, 67, 68, 69, 70, 71, 72, 73, 74, 75,
                                76, 77, 78, 80, 81, 82, 83, 84, 86, 87, 88, 89,
                                90, 91, 94, 95, 99, 100, 102, 103, 104, 106, 107, 108, 109]
            output_file = 'output.xlsx'  # Output XLSX file
            process_csv_and_save_as_xlsx(filepath, columns_to_delete, output_file)

# Register the hotkey callback function for each hotkey combination
for hotkey in keybindings.keys():
    keyboard.add_hotkey(hotkey, on_hotkey, args=(hotkey,))

# Start the keyboard listener for hotkey events
keyboard.wait('esc')  # Press 'Esc' to stop




